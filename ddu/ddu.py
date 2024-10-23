# Importações
import requests
import json
import pandas as pd
from datetime import datetime
import openpyxl
import os

# Caminhos dos arquivos
excel_file_path = 'D:/dyna/api/ddu/excel/list-ddu.xlsx'
config_file_path = 'D:/dyna/api/token/relatorio/config.json'

# Mapeamento de métricas
metric_mapping = {
    "(builtin:billing.ddu.metrics.total:splitBy():sort(value(auto,descending)):limit(20)):names": "Metrics",
    "(builtin:billing.ddu.events.total:splitBy():sort(value(auto,descending)):limit(20)):names": "Events",
    "(builtin:billing.ddu.log.total:splitBy():sort(value(auto,descending)):limit(20)):names": "Log",
    "(builtin:billing.ddu.serverless.byDescription:splitBy():sort(value(auto,descending)):limit(20)):names": "Serverless"
}

def find_next_empty_row(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if not cell.value:
                return cell.row
    return sheet.max_row + 1

# Carregar configurações
try:
    with open(config_file_path, 'r') as config_file:
        config_data = json.load(config_file)
    print("Configurações carregadas com sucesso")
except Exception as e:
    print(f"Erro ao carregar configurações: {e}")
    exit(1)

# Lista para armazenar os dados
data_to_df = []

# Período da análise (corrigido para 30 dias em setembro)
date_from, date_to = '2024-09-01', '2024-09-30'

# Iterar sobre as configurações de solicitação Dynatrace
for request_data in config_data['requests_data']:
    token = request_data['token']
    endpoint = request_data['url']
    Tenant = request_data['name']

    headers = {
        'accept': 'application/json',
        'Authorization': f'Api-Token {token}',
    }

    try:
        # Enviar solicitação para API Dynatrace
        print(f"--------- Verificando {Tenant} ---------- ")
        response = requests.get(
            f"{endpoint}/api/v2/metrics/query?metricSelector=(builtin:billing.ddu.metrics.total:splitBy():sort(value(auto,descending)):limit(20)):names,(builtin:billing.ddu.events.total:splitBy():sort(value(auto,descending)):limit(20)):names,(builtin:billing.ddu.log.total:splitBy():sort(value(auto,descending)):limit(20)):names,(builtin:billing.ddu.serverless.byDescription:splitBy():sort(value(auto,descending)):limit(20)):names&from={date_from}T00:00:00Z&to={date_to}T23:59:59Z&resolution=Inf",
            headers=headers,
        )

        if response.status_code == 200:
            try:
                response_data = json.loads(response.text)
                row_data = {
                    'Data inicio': datetime.strptime(date_from, '%Y-%m-%d').strftime('%d/%m/%Y'),
                    'Data final': datetime.strptime(date_to, '%Y-%m-%d').strftime('%d/%m/%Y'),
                    'Tenant': request_data['name']
                }

                # Processar métricas
                for metric_selector, metric_name in metric_mapping.items():
                    result_metric = next((result for result in response_data['result'] 
                                        if result['metricId'] == metric_selector), None)

                    if result_metric:
                        values_list = result_metric['data'][0]['values'] if result_metric['data'] else [0]
                        values = int(values_list[0]) if values_list else 0
                        row_data[metric_name] = values

                data_to_df.append(row_data)
                print(f"Dados processados com sucesso para {Tenant}")
            except json.JSONDecodeError as e:
                print(f"Erro ao decodificar JSON para {Tenant}: {e}")
                print("Resposta da API:", response.text)
        else:
            print(f'Erro {Tenant}: {response.status_code}')
            print("Resposta da API:", response.text)
    except Exception as e:
        print(f"Erro ao processar {Tenant}: {e}")
        continue

try:
    # Criar DataFrame
    df = pd.DataFrame(data_to_df)

    # Calcular Total DDU
    df['Total DDU'] = df[['Metrics', 'Events', 'Log', 'Serverless']].sum(axis=1)

    # Organizar colunas
    df = df[['Data inicio', 'Data final', 'Tenant', 'Metrics', 'Events', 'Log', 'Serverless', 'Total DDU']]

    # Salvar no Excel
    if os.path.isfile(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet = workbook['Sheet1']
        empty_row = find_next_empty_row(sheet)

        for _, row_data in df.iterrows():
            for col, value in enumerate(row_data.values, 1):
                sheet.cell(row=empty_row, column=col, value=value)
            empty_row += 1

        workbook.save(excel_file_path)
        print(f'Dados adicionados ao arquivo existente: {excel_file_path}')
    else:
        df.to_excel(excel_file_path, index=False)
        print(f'Novo arquivo criado: {excel_file_path}')

    print("---------- Processamento finalizado com sucesso -------------------")
except Exception as e:
    print(f"Erro ao processar ou salvar dados: {e}")