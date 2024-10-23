import os
import json
import requests
import pandas as pd
from datetime import datetime
import openpyxl

# Caminhos para os arquivos
EXCEL_FILE_PATH = 'D:/dyna/api/ddu/excel/agosto.xlsx'
CONFIG_FILE_PATH = 'D:/dyna/api/token/relatorio/config.json'

# Mapeamento de métricas originais para nomes desejados
METRIC_MAPPING = {
    "(builtin:billing.ddu.metrics.total:splitBy():sort(value(auto,descending)):limit(20)):names": "Metrics",
    "(builtin:billing.ddu.events.total:splitBy():sort(value(auto,descending)):limit(20)):names": "Events",
    "(builtin:billing.ddu.log.total:splitBy():sort(value(auto,descending)):limit(20)):names": "Log",
    "(builtin:billing.ddu.serverless.byDescription:splitBy():sort(value(auto,descending)):limit(20)):names": "Serverless"
}

def load_config(config_file_path):
    """Carrega o arquivo JSON de configuração."""
    try:
        with open(config_file_path, 'r') as config_file:
            return json.load(config_file)
    except Exception as e:
        print(f"Erro ao carregar o arquivo de configuração: {e}")
        return None

def find_next_empty_row(sheet):
    """Encontra a próxima linha vazia em uma planilha Excel."""
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if not cell.value:
                return cell.row
    return sheet.max_row + 1

def fetch_dynatrace_data(request_data, date_from, date_to):
    """Envia uma solicitação para a API da Dynatrace e retorna os dados."""
    token = request_data['token']
    endpoint = request_data['url']
    headers = {
        'accept': 'application/json',
        'Authorization': f'Api-Token {token}',
    }

    response = requests.get(
        f"{endpoint}/api/v2/metrics/query?metricSelector=(builtin:billing.ddu.metrics.total:splitBy():"
        f"sort(value(auto,descending)):limit(20)):names,(builtin:billing.ddu.events.total:splitBy():"
        f"sort(value(auto,descending)):limit(20)):names,(builtin:billing.ddu.log.total:splitBy():"
        f"sort(value(auto,descending)):limit(20)):names,(builtin:billing.ddu.serverless.byDescription:splitBy():"
        f"sort(value(auto,descending)):limit(20)):names&from={date_from}T00:00:00Z&to={date_to}T23:59:59Z&resolution=Inf",
        headers=headers,
    )

    if response.status_code == 200:
        return json.loads(response.text)
    else:
        print(f'Erro na solicitação para {request_data["name"]}: {response.status_code}')
        return None
def process_response_data(response_data, request_data):
    """Processa os dados da resposta da API da Dynatrace."""
    row_data = {
        'Nome do Cliente': request_data['name'],
        'Endpoint': request_data['url'],
        'Total DDU': 0
    }

    for metric_selector, metric_name in METRIC_MAPPING.items():
        result_metric = next((result for result in response_data['result'] if result['metricId'] == metric_selector), None)
        if result_metric:
            values_list = result_metric['data'][0]['values'] if result_metric['data'] else [0]
            values = int(values_list[0]) if values_list else 0
            row_data['Total DDU'] += values

    return row_data

def save_to_excel(df, excel_file_path):
    """Salva os dados no arquivo Excel, atualizando-o ou criando um novo."""
    if os.path.isfile(excel_file_path):
        workbook = openpyxl.load_workbook(excel_file_path)
        sheet_name = 'Sheet1'
        sheet = workbook[sheet_name]
        empty_row = find_next_empty_row(sheet)

        for _, row_data in df.iterrows():
            for col, value in enumerate(row_data.values, 1):
                sheet.cell(row=empty_row, column=col, value=value)
            empty_row += 1

        workbook.save(excel_file_path)
        print(f'Dados adicionados com sucesso no arquivo Excel: {excel_file_path}')
    else:
        df.to_excel(excel_file_path, index=False)
        print(f'Dados salvos com sucesso no arquivo Excel: {excel_file_path}')

def main():
    config_data = load_config(CONFIG_FILE_PATH)
    if config_data is None:
        return

    data_to_df = []
    date_from, date_to = '2024-09-01', '2024-09-30'

    for request_data in config_data['requests_data']:
        response_data = fetch_dynatrace_data(request_data, date_from, date_to)
        if response_data:
            row_data = process_response_data(response_data, request_data)
            data_to_df.append(row_data)

    df = pd.DataFrame(data_to_df)
    df = df[['Nome do Cliente', 'Endpoint', 'Total DDU']]
    save_to_excel(df, EXCEL_FILE_PATH)
    print("Código finalizado.")

if __name__ == "__main__":
    print("Iniciando execução do script...")
    main()
    print("Execução concluída.")
