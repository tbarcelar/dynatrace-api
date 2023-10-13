import requests
import pandas as pd
from datetime import datetime, timedelta
import json
import os
import openpyxl

def find_next_empty_row(sheet):
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=1):
        for cell in row:
            if not cell.value:
                return cell.row
    return sheet.max_row + 1

def make_api_request(url, token, start_time, end_time):
    headers = {
        'Authorization': f'Api-Token {token}',
        'accept': 'application/json; charset=utf-8',
    }
    start_time_iso = start_time.strftime('%Y-%m-%dT%H:%M:%SZ')
    end_time_iso = end_time.strftime('%Y-%m-%dT%H:%M:%SZ')
    params = {
        'from': start_time_iso,
        'to': end_time_iso,
        'includeDetails': 'true',
    }
    response = requests.get(url, params=params, headers=headers)
    if response.status_code == 200:
        return response.json()
    else:
        raise Exception(f"Erro na solicitação: {response.status_code} - {response.text}")

def main():
    file_path = 'C:/Users/tbarc/Dedalus Prime/Monitoração & Continuidade - Dynatrace/Api/consumo-ddu/excel-ddu//host-units.xlsx'
    sheet_name = 'Sheet1'
    
    with open('C:/Users/tbarc/Dedalus Prime\Monitoração & Continuidade - Automacao - Documentos/Automacao/dynatrace-api/token-scope-metrics-read,-DataExport/config.json', 'r') as f:
        config_data = json.load(f)
    
    requests_data = config_data["requests_data"]
    
    start_date = datetime(2023, 10, 1)
    end_date = datetime(2023, 10, 30)
    if os.path.exists(file_path):
        workbook = openpyxl.load_workbook(file_path)
        if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.create_sheet(sheet_name)
            sheet['A1'] = 'start_date'
            sheet['B1'] = 'end_date'
            sheet['C1'] = 'tenant'
            sheet['D1'] = 'Full_Stack'
            sheet['E1'] = 'INFRASTRUCTURE'
            sheet['F1'] = 'Total de Consumo'
    else:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = sheet_name
        sheet['A1'] = 'start_date'
        sheet['B1'] = 'end_date'
        sheet['C1'] = 'tenant'
        sheet['D1'] = 'Full_Stack'
        sheet['E1'] = 'INFRASTRUCTURE'
        sheet['F1'] = 'Total de Consumo'
    
    for request_data in requests_data:
        data = []
        url_part = '/api/v1/entity/infrastructure/hosts'
        url = request_data['url'] + url_part
        token = request_data['token']
        tenant = request_data['name']
        
        try:
            current_date = start_date
            while current_date <= end_date:
                next_date = current_date + timedelta(days=1)
                print(f"Solicitando dados para {current_date.strftime('%Y-%m-%d')} - {next_date.strftime('%Y-%m-%d')}")
                response_data = make_api_request(url, token, current_date, next_date)
                print(f"Dados da API recebidos: {response_data}")
                df = pd.DataFrame(response_data)
                
                if 'monitoringMode' in df.columns:
                    monitoringMode_infra_total = df[df['monitoringMode'] == 'INFRASTRUCTURE'].shape[0]
                    monitoringMode_full_total = df[df['monitoringMode'] == 'FULL_STACK'].shape[0]
                    data.append([current_date.strftime('%Y-%m-%d'), next_date.strftime('%Y-%m-%d'), tenant, monitoringMode_full_total, monitoringMode_infra_total, monitoringMode_full_total + monitoringMode_infra_total])
                
                current_date = next_date
        
        except Exception as e:
            print(f"Erro ao processar a solicitação para {tenant}: {e}")
            data.append([current_date.strftime('%Y-%m-%d'), next_date.strftime('%Y-%m-%d'), tenant, 0, 0, 0])
        
        for row_data in data:
            empty_row = find_next_empty_row(sheet)
            for col, value in enumerate(row_data, 1):
                sheet.cell(row=empty_row, column=col, value=value)
    
    workbook.save(file_path)
    print("API processada com sucesso.")
    print(f"Dados salvos em '{file_path}'.")

if __name__ == "__main__":
    main()
