import requests
import json
import openpyxl

# Caminhos de arquivo
config_file_path = "D:/dyna/api/tenant/config.json"
excel_file_path = "D:/dyna/api/tenant/verificar-tenant/excel/mz.xlsx"

# Função para obter a URL e o token do JSON
def obter_configuracao():
    try:
        with open(config_file_path, "r") as config_file:
            config_data = json.load(config_file)
        url = config_data["requests_data"][0]["url"]
        token = config_data["requests_data"][0]["token"]
        return url, token
    except Exception as e:
        print(f"Erro ao carregar a configuração: {e}")
        raise

# Função principal para executar o processo
def main():
    try:
        url, token = obter_configuracao()
        api_url = f"{url}/api/config/v1/managementZones"
        headers = {
            'Accept': 'application/json; charset=utf-8',
            'Content-Type': 'application/json; charset=utf-8',
            "Authorization": f"Api-Token {token}"
        }
        params = {}
        response = requests.get(api_url, params=params, headers=headers)
        if response.status_code == 200:
            data = response.json()["values"]
            process_data(data)
        else:
            print(f"Falha na requisição: {response.status_code} - {response.text}")
    except Exception as e:
        print(f"Erro ao executar o processo: {e}")

# Função auxiliar para processar os dados da API
def process_data(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.cell(row=1, column=1, value="id")
    ws.cell(row=1, column=2, value="name")
    for idx, item in enumerate(data, start=2):
        ws.cell(row=idx, column=1, value=item["id"])
        ws.cell(row=idx, column=2, value=item["name"])
    wb.save(excel_file_path)
    print("------------------  MZ finalizado --------------------------")
    print(f"Arquivo Excel salvo com sucesso em: {excel_file_path}")

if __name__ == "__main__":
    print("Iniciando o processo...")
    main()
    print("Processo concluído.")
