import json
import requests
import openpyxl

# Carrega o arquivo config.json
with open("C:/Users/config.json", "r") as config_file:
    config_data = json.load(config_file)

# Obtém a URL e o token do JSON
url = config_data["requests_data"][0]["url"]
token = config_data["requests_data"][0]["token"]

# Define a URL da API Dynatrace
api_url = f"{url}/api/config/v1/managementZones"

# Define os cabeçalhos com o token
headers = {
    'Accept': 'application/json; charset=utf-8',
    'Content-Type': 'application/json; charset=utf-8',
    "Authorization": f"Api-Token {token}",
}

params = {}

# Faz a solicitação GET para a API
response = requests.get(api_url, params=params, headers=headers)

# Verifica se a resposta foi bem-sucedida
if response.status_code == 200:
    data = response.json()["values"]

    # Cria um novo arquivo Excel
    wb = openpyxl.Workbook()
    ws = wb.active

    # Adiciona cabeçalhos às colunas
    ws.cell(row=1, column=1, value="id")
    ws.cell(row=1, column=2, value="name")

    # Preenche os dados nas colunas
    for idx, item in enumerate(data, start=2):
        ws.cell(row=idx, column=1, value=item["id"])
        ws.cell(row=idx, column=2, value=item["name"])

    # Salva o arquivo Excel
    wb.save("C:/Users/mz.xlsx")

    print("Dados salvos com sucesso ")
else:
    print(f"Falha na solicitação API, status code: {response.status_code}")
